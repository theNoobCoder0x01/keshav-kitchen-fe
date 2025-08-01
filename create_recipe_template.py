import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_recipe_template():
    # Sample data for the template
    sample_data = [
        {
            "Recipe Name": "Chicken Curry",
            "Category": "Main Course",
            "Subcategory": "Indian",
            "Description": "A delicious chicken curry with aromatic spices",
            "Instructions": "1. Marinate chicken with spices\n2. Heat oil in a pan\n3. Add onions and cook until golden\n4. Add chicken and cook until done\n5. Add tomatoes and simmer",
            "Servings": 4,
            "Ingredients": "Chicken,500,g,0.02;Onion,100,g,0.01;Tomato,150,g,0.015;Spices,50,g,0.05;Oil,30,ml,0.02"
        },
        {
            "Recipe Name": "Rice Pilaf",
            "Category": "Side Dish",
            "Subcategory": "Indian",
            "Description": "Fragrant basmati rice cooked with spices",
            "Instructions": "1. Wash and soak rice for 30 minutes\n2. Heat oil and add whole spices\n3. Add rice and water\n4. Cook until done",
            "Servings": 6,
            "Ingredients": "Basmati Rice,300,g,0.03;Cardamom,5,pcs,0.1;Cinnamon,1,stick,0.05;Oil,20,ml,0.02"
        },
        {
            "Recipe Name": "Dal Makhani",
            "Category": "Main Course",
            "Subcategory": "Indian",
            "Description": "Creamy black lentils cooked overnight",
            "Instructions": "1. Soak lentils overnight\n2. Cook in pressure cooker\n3. Add cream and butter\n4. Simmer until creamy",
            "Servings": 8,
            "Ingredients": "Black Lentils,200,g,0.025;Kidney Beans,100,g,0.02;Cream,100,ml,0.03;Butter,50,g,0.04;Spices,30,g,0.05"
        },
        {
            "Recipe Name": "Naan Bread",
            "Category": "Bread",
            "Subcategory": "Indian",
            "Description": "Soft and fluffy Indian flatbread",
            "Instructions": "1. Mix flour, yeast, and water\n2. Knead for 10 minutes\n3. Let rise for 2 hours\n4. Shape and cook on hot griddle",
            "Servings": 10,
            "Ingredients": "All Purpose Flour,500,g,0.015;Yeast,10,g,0.05;Yogurt,100,ml,0.02;Oil,30,ml,0.02"
        },
        {
            "Recipe Name": "Gulab Jamun",
            "Category": "Dessert",
            "Subcategory": "Indian",
            "Description": "Sweet milk dumplings in sugar syrup",
            "Instructions": "1. Mix milk powder and flour\n2. Add ghee and make dough\n3. Shape into balls\n4. Fry until golden\n5. Soak in sugar syrup",
            "Servings": 12,
            "Ingredients": "Milk Powder,200,g,0.04;Flour,50,g,0.015;Ghee,100,g,0.06;Sugar,300,g,0.02;Cardamom,10,pcs,0.1"
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(sample_data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Recipe Import Template"
    
    # Add headers
    headers = ["Recipe Name", "Category", "Subcategory", "Description", "Instructions", "Servings", "Ingredients"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, (key, value) in enumerate(row_data.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Recipe Import Template - Instructions"],
        [""],
        ["Column Descriptions:"],
        ["Recipe Name", "Required. The name of the recipe"],
        ["Category", "Required. Main category (e.g., Main Course, Dessert, etc.)"],
        ["Subcategory", "Required. Subcategory (e.g., Indian, Italian, etc.)"],
        ["Description", "Optional. Brief description of the recipe"],
        ["Instructions", "Optional. Step-by-step cooking instructions"],
        ["Servings", "Optional. Number of servings (whole number)"],
        ["Ingredients", "Optional. Ingredients in format: Name,Quantity,Unit,CostPerUnit;Name2,Quantity2,Unit2,CostPerUnit2"],
        [""],
        ["Important Notes:"],
        ["- First row contains headers. Do not modify these."],
        ["- Recipe Name, Category, and Subcategory are required fields."],
        ["- Ingredients format: 'Chicken,500,g,0.02;Onion,100,g,0.01'"],
        ["- Each ingredient should have: Name, Quantity, Unit, CostPerUnit (optional)"],
        ["- Separate multiple ingredients with semicolon (;)"],
        ["- CostPerUnit is optional for each ingredient"],
        ["- Servings should be a whole number"],
        ["- Instructions can include line breaks using \\n"]
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 3 or row_idx == 12:
                cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    wb.save("recipe_import_template.xlsx")
    print("Recipe import template created successfully!")
    print("File: recipe_import_template.xlsx")
    print("\nTemplate includes:")
    print("- Sample recipes with proper formatting")
    print("- Instructions sheet with detailed guidelines")
    print("- Proper column headers and data types")

if __name__ == "__main__":
    create_recipe_template()